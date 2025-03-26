from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook("report_data.xlsx")
ws = wb.active

fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

header = [cell.value for cell in ws[1]]
status_index = header.index("Status") + 1

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    cell = row[status_index - 1]
    if cell.value == "Failed":
        for c in row:
            c.fill = fill

wb.save("report_data3.xlsx")